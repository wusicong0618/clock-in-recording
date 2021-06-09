#!/usr/bin/python
# -*- coding: UTF-8 -*-

import xlrd,xlwt
from xlutils.copy import copy
import time,datetime
from datetime import date
from datetime import timedelta
import win32com.client as win32
import os,sys
import re
import openpyxl

Mywb=xlwt.Workbook()
Mysheet=Mywb.add_sheet('sheet1')
Wkno,Belongdate,Startdate,Starttime,Enddate,Endtime,Overtime=[],[],[],[],[],[],[]
OvertimeType,Name=[],[]

#寻找文件
print('寻找文件...................................')
m=0
path=os.getcwd()
print("当前路径"+path)
for dirpath, dirnames, filenames in os.walk(path):
    for filename in filenames:
        if '号加班' in filename:
            FOLname=filename
            print('复制'+FOLname)
            m=m+1
        elif '号白班加班' in filename:
            EOLdayname=filename
            print('复制'+EOLdayname)
            m=m+1
        elif '号夜班加班' in filename:
            EOLnightname=filename
            print('复制'+EOLnightname)
            m=m+1
        elif '打卡' in filename:
            checkrecord=filename
if m<3:
    print('文件不全')
    sys.exit()

#复制前段加班内容
print('复制前段加班内容...................................')
wb1=xlrd.open_workbook(FOLname)
sheet1=wb1.sheet_by_index(0)
row1=sheet1.nrows
column1=sheet1.ncols
n1=0
for i1 in range(0,row1):
    Wkdatebelong1=sheet1.cell_value(i1,5)
    if Wkdatebelong1=='':
        n1=n1+1
    else:
        for j1 in range(0,column1):
            content1=sheet1.cell_value(i1,j1)
            Mysheet.write(i1-n1,j1,content1)
        if i1>0:
            Wkno.append(sheet1.cell_value(i1,0))
            Belongdate.append(sheet1.cell_value(i1,5))
            Startdate.append(sheet1.cell_value(i1,6))
            Starttime.append(sheet1.cell_value(i1,7))
            Enddate.append(sheet1.cell_value(i1,8))
            Endtime.append(sheet1.cell_value(i1,9))
            Overtime.append(sheet1.cell_value(i1,13))
            OvertimeType.append(sheet1.cell_value(i1,4))
            Name.append(sheet1.cell_value(i1,1))
            
#复制后段白班加班内容
print('复制后段白班加班内容...................................')
wb2=xlrd.open_workbook(EOLdayname)
sheet2=wb2.sheet_by_index(0)
row2=sheet2.nrows
column2=sheet2.ncols
n2=0
for i2 in range(1,row2):
    Wkdatebelong2=sheet2.cell_value(i2,5)
    if Wkdatebelong2=='':
        n2=n2+1
    else:
        for j2 in range(0,column2):
            content2=sheet2.cell_value(i2,j2)
            Mysheet.write(i2-n2+row1-n1-1,j2,content2)
        Wkno.append(sheet2.cell_value(i2,0))
        Belongdate.append(sheet2.cell_value(i2,5))
        Startdate.append(sheet2.cell_value(i2,6))
        Starttime.append(sheet2.cell_value(i2,7))
        Enddate.append(sheet2.cell_value(i2,8))
        Endtime.append(sheet2.cell_value(i2,9))
        Overtime.append(sheet2.cell_value(i2,13))
        OvertimeType.append(sheet2.cell_value(i2,4))
        Name.append(sheet2.cell_value(i2,1))
        
#复制后段夜班加班内容
print('复制后段夜班加班内容...................................')
wb3=xlrd.open_workbook(EOLnightname)
sheet3=wb3.sheet_by_index(0)
row3=sheet3.nrows
column3=sheet3.ncols
n3=0
for i3 in range(1,row3):
    Wkdatebelong3=sheet3.cell_value(i3,5)
    if Wkdatebelong3=='':
        n3=n3+1
    else:
        for j3 in range(0,column3):
            content3=sheet3.cell_value(i3,j3)
            Mysheet.write(i3-n3+row2-n2+row1-n1-2,j3,content3)
        Wkno.append(sheet3.cell_value(i3,0))
        Belongdate.append(sheet3.cell_value(i3,5))
        Startdate.append(sheet3.cell_value(i3,6))
        Starttime.append(sheet3.cell_value(i3,7))
        Enddate.append(sheet3.cell_value(i3,8))
        Endtime.append(sheet3.cell_value(i3,9))
        Overtime.append(sheet3.cell_value(i3,13))
        OvertimeType.append(sheet3.cell_value(i3,4))
        Name.append(sheet3.cell_value(i3,1))

#判断文件名
print('判断文件名...................................')
pattern = re.compile(r'[\u4e00-\u9fa5]')
FOLunchinese = re.sub(pattern,"",FOLname).replace(" ","").replace(".xls","")
EOLdayunchinese = re.sub(pattern,"",EOLdayname).replace(" ","").replace(".xls","")
EOLnightunchinese = re.sub(pattern,"",EOLnightname).replace(" ","").replace(".xls","")

if FOLunchinese==EOLdayunchinese and EOLdayunchinese==EOLnightunchinese:
    Newname=FOLunchinese+'加班前后段汇总.xls'
    print('日期'+FOLunchinese)
else:
    print('文件名有误')
    sys.exit()

#保存
Mywb.save(Newname)
print('合并完成')

MywbRead=xlrd.open_workbook(Newname)
MysheetRead=MywbRead.sheet_by_index(0)
MywbWrite=copy(MywbRead)
MysheetWrite=MywbWrite.get_sheet(0)
Myrow=MysheetRead.nrows

#判断工号是否有重复
print('判断工号是否有重复...................................')
Checkno=[]
for c0 in Wkno:
    if c0 not in Checkno:
        Checkno.append(c0)
if len(Wkno)==len(Checkno):
    print('工号无重复')
else:
    print('工号重复')

#判断加班类型
print('判断加班类型...................................')
for c4 in range(0,len(OvertimeType)):
    if '员工' not in str(OvertimeType[c4]):
        MysheetWrite.write(c4+1,4,'员工'+OvertimeType[c4])
        print (Name[c4]+'修改加班类型')


#判断用餐时间扣除

#修改归属日期
print('修改归属日期...................................')
for c5 in range(0,len(Belongdate)):
    if type(Belongdate[c5]) is float:
        tuplec5=xlrd.xldate_as_tuple(Belongdate[c5],0)
        Newc5 = date(*tuplec5[0:3]).strftime('%Y/%m/%d')
        MysheetWrite.write(c5+1,5,Newc5)
        print(Name[c5]+'修改归属日期格式')
    else:
        MysheetWrite.write(c5+1,5,Belongdate[c5])


#修改开始日期
print('修改开始日期...................................')
NewStartdate=[]
for c6 in range(0,len(Startdate)):
    if type(Startdate[c6]) is float:
        tuplec6=xlrd.xldate_as_tuple(Startdate[c6],0)
        Newc6 = date(*tuplec6[0:3]).strftime('%Y/%m/%d')
        MysheetWrite.write(c6+1,6,Newc6)
        print(Name[c6]+'修改开始日期格式')
        NewStartdate.append(Newc6)
    else:
        MysheetWrite.write(c6+1,6,Startdate[c6])
        NewStartdate.append(Startdate[c6])

#修改开始时间
print('修改开始时间...................................')
NewStarttime=[]
for c7 in range(0,len(Starttime)):
    if type(Starttime[c7]) is float:
        tuplec7=xlrd.xldate.xldate_as_datetime(Starttime[c7],0)
        Newc7=tuplec7.strftime('%H:%M')
        #print(Newc7)
        MysheetWrite.write(c7+1,7,Newc7)
        print(Name[c7]+'修改开始时间格式')
        NewStarttime.append(Newc7)
    else:
        MysheetWrite.write(c7+1,7,Starttime[c7])
        NewStarttime.append(Starttime[c7])

#修改结束日期
print('修改结束日期...................................')
NewEnddate=[]
for c8 in range(0,len(Enddate)):
    if type(Enddate[c8]) is float:
        tuplec8=xlrd.xldate_as_tuple(Enddate[c8],0)
        Newc8=date(*tuplec8[0:3]).strftime('%Y/%m/%d')
        MysheetWrite.write(c8+1,8,Newc8)
        print(Name[c8]+'修改结束日期格式')
        NewEnddate.append(Newc8)
    else:
        MysheetWrite.write(c8+1,8,Enddate[c8])
        NewEnddate.append(Enddate[c8])

#修改结束时间
print('修改结束时间...................................')
NewEndtime=[]
for c9 in range(0,len(Endtime)):
    if type(Endtime[c9]) is float:
        tuplec9=xlrd.xldate.xldate_as_datetime(Endtime[c9],0)
        Newc9=tuplec9.strftime('%H:%M')
        #print(Newc9)
        MysheetWrite.write(c9+1,9,Newc9)
        print(Name[c9]+'修改结束时间格式')
        NewEndtime.append(Newc9)
    else:
        MysheetWrite.write(c9+1,9,Endtime[c9])
        NewEndtime.append(Endtime[c9])

#判断加班时间
print('判断加班时间...................................')
for c13 in range(0,len(Overtime)):
    Endtimecal=datetime.datetime.strptime(str(str(NewEnddate[c13])+str(NewEndtime[c13])),'%Y/%m/%d%H:%M')
    Starttimecal=datetime.datetime.strptime(str(str(NewStartdate[c13])+str(NewStarttime[c13])),'%Y/%m/%d%H:%M')
    Actualtime=((Endtimecal-Starttimecal).total_seconds())/3600
    if (Endtimecal-Starttimecal).days<0:
        print(Name[c13]+"时间异常")
    else:
        if Actualtime!=Overtime[c13]:
            print(Name[c13]+"加班时间错误")


#根据考勤记录核对加班时间
print('根据考勤记录核对加班时间...................................')
wb4 = openpyxl.load_workbook(checkrecord)
sheet4=wb4.active
row4=sheet4.max_row
column4=sheet4.max_column

if '.' in FOLunchinese:
    Ydate=FOLunchinese.replace('.','-')
for k in range(0,len(Wkno)):
    for i4 in range(1,row4+1):
        Wknocheck=sheet4.cell(row=i4,column=1).value
        if Wkno[k]==Wknocheck:
            for j4 in range(1,column4+1):
                Date=sheet4.cell(row=2,column=j4).value
                if Ydate in str(Date):
                    GoodYdate=datetime.datetime.strptime(Ydate,'%m-%d')+timedelta(days=121*365+30)
                    GoodEnddate=datetime.datetime.strptime(str(NewEnddate[k]),'%Y/%m/%d')
                    if GoodEnddate==GoodYdate:
                        Checktime=sheet4.cell(row=i4,column=j4).value
                        NewChecktime=Ydate+str(Checktime)[-5:]
                        TimeChecktime=datetime.datetime.strptime(NewChecktime,'%m-%d%H:%M')+timedelta(days=121*365+30)
                    else:
                        Checktime=sheet4.cell(row=i4,column=j4+1).value
                        NewChecktime=Ydate+str(Checktime)[-5:]
                        TimeChecktime=datetime.datetime.strptime(NewChecktime,'%m-%d%H:%M')+timedelta(days=121*365+31)
                    TimeEndtime=datetime.datetime.strptime(str(str(NewEnddate[k])+str(NewEndtime[k])),'%Y/%m/%d%H:%M')
                    if TimeEndtime>(TimeChecktime):
                        print(Name[k]+'考勤时间异常')
                        print("加班时间"+str(TimeEndtime))
                        print("打卡时间"+str(TimeChecktime))
                    

#保存
MywbWrite.save(Newname)
fname = path+"\\"+Newname
print("保存路径"+fname)

#转换格式
excel = win32.DispatchEx('Excel.Application')
wb = excel.Workbooks.Open(fname)
wb.SaveAs(fname+"x", FileFormat = 51)
wb.Close()
excel.Application.Quit()
print('转换xlsx格式')


