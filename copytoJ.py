#!/usr/bin/python
# -*- coding: UTF-8 -*-

import xlrd
import os,sys
import openpyxl

#寻找文件
m=0
path=os.getcwd()
print("当前路径"+path)
for dirpath, dirnames, filenames in os.walk(path):
    for filename in filenames:
        if '号加班' in filename:
            FOLname=filename
            m=m+1
        elif '号白班加班' in filename:
            EOLdayname=filename
            m=m+1
        elif '号夜班加班' in filename:
            EOLnightname=filename
            m=m+1
        elif '加班前后段汇总.xlsx' in filename:
            OriginFile=filename
if m<3:
    print('文件不全')
    sys.exit()

#读取前段人员工号
FOLWkno=[]
wb1=xlrd.open_workbook(FOLname)
sheet1=wb1.sheet_by_index(0)
row1=sheet1.nrows
for i1 in range(0,row1):
    FOLWkno.append(sheet1.cell_value(i1,0))
print('读取'+FOLname)

#读取后段白班人员工号
EOLdayWkno=[]
wb2=xlrd.open_workbook(EOLdayname)
sheet2=wb2.sheet_by_index(0)
row2=sheet2.nrows
for i2 in range(0,row2):
    EOLdayWkno.append(sheet2.cell_value(i2,0))
print('读取'+EOLdayname)

#读取后段夜班人员工号
EOLnightWkno=[]
wb3=xlrd.open_workbook(EOLnightname)
sheet3=wb3.sheet_by_index(0)
row3=sheet3.nrows
for i3 in range(0,row3):
    EOLnightWkno.append(sheet3.cell_value(i3,0))
print('读取'+EOLnightname)

#读取原文件
wb4 = openpyxl.load_workbook(OriginFile)
sheet4=wb4.active
row4=sheet4.max_row
column4=sheet4.max_column

#新建
Mywb=openpyxl.Workbook()
Mysheet=Mywb.active

#复制原文件
for n in range(1,row4+1):
    for m in range(1,column4+1):
        copycell=sheet4.cell(row=n,column=m).value
        Mysheet.cell(row=n, column=m).value=copycell
print('复制'+OriginFile)

#插入2列
Mysheet.insert_cols(3,2)

#写入
for i4 in range(1,row4+1):
    Wknocheck=sheet4.cell(row=i4,column=1).value
    if Wknocheck=='工号':
        Mysheet.cell(row=i4, column=3).value='段位'
        Mysheet.cell(row=i4, column=4).value='班次'
        print('写入标题')
    else:
        for k1 in range(0,len(FOLWkno)):
            if Wknocheck==FOLWkno[k1]:
                Mysheet.cell(row=i4, column=3).value='前段'
                Mysheet.cell(row=i4, column=4).value='白班'
        for k2 in range(0,len(EOLdayWkno)):
            if Wknocheck==EOLdayWkno[k2]:
                Mysheet.cell(row=i4, column=3).value='后段'
                Mysheet.cell(row=i4, column=4).value='白班'
        for k3 in range(0,len(EOLnightWkno)):
            if Wknocheck==EOLnightWkno[k3]:
                Mysheet.cell(row=i4, column=3).value='后段'
                Mysheet.cell(row=i4, column=4).value='夜班'        
print('写入段位/班次')
        
#保存
Mywb.save('To '+OriginFile)




